"""
Экспорт результатов в Excel (ПОЛНЫЙ КОД)
"""
from typing import List, Dict, Optional
from datetime import datetime
import os

from utils.logger import get_logger
from storage.models import KeywordData

logger = get_logger('WordStat.Exporter')

class ExcelExporter:
    """Экспорт результатов в Excel"""
    
    def __init__(self):
        """Инициализация"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            self.openpyxl = openpyxl
            self.Font = Font
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.Border = Border
            self.Side = Side
            logger.info("✓ ExcelExporter инициализирован (openpyxl)")
        except ImportError:
            logger.error("✗ openpyxl не установлен")
            raise
        
        self.wb = None
    
    def _format_header_cell(self, cell, text: str) -> None:
        """Форматировать ячейку заголовка"""
        cell.value = text
        cell.font = self.Font(bold=True, color="FFFFFF", size=12)
        cell.fill = self.PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.alignment = self.Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _format_data_cell(self, cell, value) -> None:
        """Форматировать ячейку д��нных"""
        cell.value = value
        cell.alignment = self.Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = self.Border(
            left=self.Side(style='thin'),
            right=self.Side(style='thin'),
            top=self.Side(style='thin'),
            bottom=self.Side(style='thin')
        )
    
    def export_seo_core(self,
                       keywords: Dict[str, KeywordData],
                       output_path: str = "output_seo_core.xlsx") -> bool:
        """
        Экспорт SEO-ядра
        """
        try:
            if not isinstance(keywords, dict):
                logger.error(f"✗ keywords должен быть dict")
                return False
            
            if not output_path or not isinstance(output_path, str):
                logger.error(f"✗ output_path некорректен")
                return False
            
            self.wb = self.openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            
            # ЛИСТ 1 - Все ключи
            ws_all = self.wb.create_sheet("Все ключи")
            headers = ["Фраза", "Count", "Глубина", "Seed", "Intent", "Timestamp"]
            ws_all.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_all.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            try:
                sorted_kw = sorted(keywords.items(), key=lambda x: x[1].count, reverse=True)
                
                for row_idx, (phrase, kwd) in enumerate(sorted_kw, 2):
                    if not isinstance(kwd, KeywordData):
                        logger.warning(f"⚠ Некорректный KeywordData: {phrase}")
                        continue
                    
                    ws_all.cell(row=row_idx, column=1).value = kwd.phrase
                    ws_all.cell(row=row_idx, column=2).value = int(kwd.count)
                    ws_all.cell(row=row_idx, column=3).value = int(kwd.depth)
                    ws_all.cell(row=row_idx, column=4).value = kwd.seed
                    ws_all.cell(row=row_idx, column=5).value = kwd.intent or "общий"
                    ws_all.cell(row=row_idx, column=6).value = kwd.timestamp
            
            except Exception as e:
                logger.error(f"✗ Ошибка добавления данных: {e}")
                return False
            
            ws_all.column_dimensions['A'].width = 45
            ws_all.column_dimensions['B'].width = 15
            ws_all.column_dimensions['C'].width = 12
            ws_all.column_dimensions['D'].width = 35
            ws_all.column_dimensions['E'].width = 18
            ws_all.column_dimensions['F'].width = 22
            
            # ЛИСТ 2 - Settings
            ws_settings = self.wb.create_sheet("Настройки")
            
            settings_data = [
                ["Параметр", "Значение"],
                ["Дата экспорта", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Всего ключей", len(keywords)],
                ["Максимальный count", max([k.count for k in keywords.values()]) if keywords else 0],
                ["Минимальный count", min([k.count for k in keywords.values()]) if keywords else 0],
                ["Среднее count", int(sum([k.count for k in keywords.values()]) / len(keywords)) if keywords else 0],
            ]
            
            for row_data in settings_data:
                ws_settings.append(row_data)
            
            for col_idx, header in enumerate(["Параметр", "Значение"], 1):
                cell = ws_settings.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            ws_settings.column_dimensions['A'].width = 30
            ws_settings.column_dimensions['B'].width = 45
            
            try:
                self.wb.save(output_path)
                logger.info(f"✓ SEO-ядро экспортировано: {output_path}")
                return True
            
            except Exception as e:
                logger.error(f"✗ Ошибка сохранения Excel: {e}")
                return False
        
        except Exception as e:
            logger.error(f"✗ Критическая ошибка export_seo_core: {e}")
            return False
    
    def export_ppc_context(self,
                          keywords: Dict[str, KeywordData],
                          output_path: str = "output_ppc_context.xlsx") -> bool:
        """
        Экспорт Контекста (PPC)
        """
        try:
            if not isinstance(keywords, dict):
                logger.error(f"✗ keywords должен быть dict")
                return False
            
            self.wb = self.openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            
            # ЛИСТ 1: Ключевые слова для PPC
            ws_keywords = self.wb.create_sheet("Ключевые слова")
            headers = ["Фраза", "Count", "Intent", "Тип ставки", "Рекомендация"]
            ws_keywords.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_keywords.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            sorted_kw = sorted(keywords.items(), key=lambda x: x[1].count, reverse=True)
            
            for row_idx, (phrase, kwd) in enumerate(sorted_kw, 2):
                if not isinstance(kwd, KeywordData):
                    continue
                
                intent = kwd.intent or "общий"
                
                if intent == "commercial":
                    bid_type = "Высокая"
                    recommendation = "Обязательно включить"
                elif intent == "info":
                    bid_type = "Средняя"
                    recommendation = "Можно включить"
                else:
                    bid_type = "Низкая"
                    recommendation = "По усмотрению"
                
                ws_keywords.cell(row=row_idx, column=1).value = kwd.phrase
                ws_keywords.cell(row=row_idx, column=2).value = int(kwd.count)
                ws_keywords.cell(row=row_idx, column=3).value = intent
                ws_keywords.cell(row=row_idx, column=4).value = bid_type
                ws_keywords.cell(row=row_idx, column=5).value = recommendation
            
            ws_keywords.column_dimensions['A'].width = 45
            ws_keywords.column_dimensions['B'].width = 15
            ws_keywords.column_dimensions['C'].width = 18
            ws_keywords.column_dimensions['D'].width = 18
            ws_keywords.column_dimensions['E'].width = 25
            
            # ЛИСТ 2: Минус-слова
            ws_negatives = self.wb.create_sheet("Минус-слова")
            headers = ["Минус-фраза", "Тип защиты"]
            ws_negatives.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_negatives.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            suggested_negatives = [
                ["бесплатно", "Защита от дешевых"],
                ["скачать", "Качество трафика"],
                ["пиратский", "Защита бренда"],
                ["demo", "Защита версии"],
            ]
            
            for row_idx, (neg_phrase, neg_type) in enumerate(suggested_negatives, 2):
                ws_negatives.cell(row=row_idx, column=1).value = neg_phrase
                ws_negatives.cell(row=row_idx, column=2).value = neg_type
            
            ws_negatives.column_dimensions['A'].width = 45
            ws_negatives.column_dimensions['B'].width = 30
            
            try:
                self.wb.save(output_path)
                logger.info(f"✓ Контекст (PPC) экспортирован: {output_path}")
                return True
            except Exception as e:
                logger.error(f"✗ Ошибка сохранения PPC Excel: {e}")
                return False
        
        except Exception as e:
            logger.error(f"✗ Критическая ошибка export_ppc_context: {e}")
            return False
    
    def export_content_plan(self,
                           keywords: Dict[str, KeywordData],
                           output_path: str = "output_content_plan.xlsx") -> bool:
        """
        Экспорт Контент-плана
        """
        try:
            if not isinstance(keywords, dict):
                logger.error(f"✗ keywords должен быть dict")
                return False
            
            self.wb = self.openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            
            # ЛИСТ 1: Темы для контента
            ws_topics = self.wb.create_sheet("Темы контента")
            headers = ["Название статьи", "Приоритет", "Примерный объем", "Keywords", "Intent"]
            ws_topics.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_topics.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            sorted_kw = sorted(keywords.items(), key=lambda x: x[1].count, reverse=True)
            
            for row_idx, (phrase, kwd) in enumerate(sorted_kw[:20], 2):
                if not isinstance(kwd, KeywordData):
                    continue
                
                if kwd.count > 1000:
                    priority = "Высокий"
                    volume = "3000-5000 слов"
                elif kwd.count > 100:
                    priority = "Средний"
                    volume = "1500-3000 слов"
                else:
                    priority = "Низкий"
                    volume = "500-1500 слов"
                
                ws_topics.cell(row=row_idx, column=1).value = f"Статья: {kwd.phrase}"
                ws_topics.cell(row=row_idx, column=2).value = priority
                ws_topics.cell(row=row_idx, column=3).value = volume
                ws_topics.cell(row=row_idx, column=4).value = kwd.phrase
                ws_topics.cell(row=row_idx, column=5).value = kwd.intent or "общий"
            
            ws_topics.column_dimensions['A'].width = 50
            ws_topics.column_dimensions['B'].width = 15
            ws_topics.column_dimensions['C'].width = 20
            ws_topics.column_dimensions['D'].width = 45
            ws_topics.column_dimensions['E'].width = 18
            
            # ЛИСТ 2: Все ключевые слова
            ws_all_kw = self.wb.create_sheet("Ключевые слова")
            headers = ["Фраза", "Count", "Тип контента", "Intent", "Глубина"]
            ws_all_kw.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_all_kw.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            for row_idx, (phrase, kwd) in enumerate(sorted_kw, 2):
                if not isinstance(kwd, KeywordData):
                    continue
                
                word_count = len(kwd.phrase.split())
                if word_count >= 4:
                    content_type = "Long-form"
                elif word_count >= 2:
                    content_type = "Medium"
                else:
                    content_type = "Short-form"
                
                ws_all_kw.cell(row=row_idx, column=1).value = kwd.phrase
                ws_all_kw.cell(row=row_idx, column=2).value = int(kwd.count)
                ws_all_kw.cell(row=row_idx, column=3).value = content_type
                ws_all_kw.cell(row=row_idx, column=4).value = kwd.intent or "общий"
                ws_all_kw.cell(row=row_idx, column=5).value = int(kwd.depth)
            
            ws_all_kw.column_dimensions['A'].width = 45
            ws_all_kw.column_dimensions['B'].width = 15
            ws_all_kw.column_dimensions['C'].width = 18
            ws_all_kw.column_dimensions['D'].width = 18
            ws_all_kw.column_dimensions['E'].width = 12
            
            try:
                self.wb.save(output_path)
                logger.info(f"✓ Контент-план экспортирован: {output_path}")
                return True
            except Exception as e:
                logger.error(f"✗ Ошибка сохранения контент-плана: {e}")
                return False
        
        except Exception as e:
            logger.error(f"✗ Критическая ошибка export_content_plan: {e}")
            return False
    
    def export_ai_clusters(self,
                          clusters: Dict[str, List[str]],
                          keywords: Optional[Dict[str, 'KeywordData']] = None,
                          output_path: str = "output_ai_clusters.xlsx") -> bool:
        """
        Экспорт AI кластеров
        
        Args:
            clusters: Словарь кластеров {центр: [фразы]}
            keywords: Опциональный словарь с данными ключевых слов
            output_path: Путь для сохранения
        """
        try:
            if not clusters:
                logger.warning("⚠ Нет кластеров для экспорта")
                return False
            
            self.wb = self.openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            
            # ЛИСТ 1: Все кластеры
            ws_clusters = self.wb.create_sheet("Кластеры")
            headers = ["Кластер", "Фраза", "Count", "Глубина"]
            ws_clusters.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_clusters.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            row_idx = 2
            for cluster_name, phrases in clusters.items():
                for phrase in phrases:
                    ws_clusters.cell(row=row_idx, column=1).value = cluster_name
                    ws_clusters.cell(row=row_idx, column=2).value = phrase
                    
                    # Добавить данные из keywords если есть
                    if keywords and phrase in keywords:
                        kwd = keywords[phrase]
                        ws_clusters.cell(row=row_idx, column=3).value = int(kwd.count)
                        ws_clusters.cell(row=row_idx, column=4).value = int(kwd.depth)
                    else:
                        ws_clusters.cell(row=row_idx, column=3).value = ""
                        ws_clusters.cell(row=row_idx, column=4).value = ""
                    
                    row_idx += 1
            
            ws_clusters.column_dimensions['A'].width = 45
            ws_clusters.column_dimensions['B'].width = 45
            ws_clusters.column_dimensions['C'].width = 15
            ws_clusters.column_dimensions['D'].width = 12
            
            # ЛИСТ 2: Сводка по кластерам
            ws_summary = self.wb.create_sheet("Сводка")
            summary_headers = ["Кластер", "Количество фраз", "% от общего"]
            ws_summary.append(summary_headers)
            
            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            total_phrases = sum(len(phrases) for phrases in clusters.values())
            row_idx = 2
            
            for cluster_name, phrases in sorted(clusters.items(), key=lambda x: len(x[1]), reverse=True):
                ws_summary.cell(row=row_idx, column=1).value = cluster_name
                ws_summary.cell(row=row_idx, column=2).value = len(phrases)
                percentage = (len(phrases) / total_phrases * 100) if total_phrases > 0 else 0
                ws_summary.cell(row=row_idx, column=3).value = f"{percentage:.1f}%"
                row_idx += 1
            
            # Итого
            ws_summary.cell(row=row_idx, column=1).value = "ИТОГО"
            ws_summary.cell(row=row_idx, column=2).value = total_phrases
            ws_summary.cell(row=row_idx, column=3).value = "100%"
            
            ws_summary.column_dimensions['A'].width = 45
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 15
            
            # ЛИСТ 3: Настройки экспорта
            ws_settings = self.wb.create_sheet("Настройки")
            settings_data = [
                ["Параметр", "Значение"],
                ["Дата экспорта", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Всего кластеров", len(clusters)],
                ["Всего фраз", total_phrases],
                ["Средний размер кластера", round(total_phrases / len(clusters), 1) if clusters else 0],
            ]
            
            for row_data in settings_data:
                ws_settings.append(row_data)
            
            for col_idx, header in enumerate(["Параметр", "Значение"], 1):
                cell = ws_settings.cell(row=1, column=col_idx)
                self._format_header_cell(cell, header)
            
            ws_settings.column_dimensions['A'].width = 30
            ws_settings.column_dimensions['B'].width = 45
            
            try:
                self.wb.save(output_path)
                logger.info(f"✓ AI кластеры экспортированы: {output_path}")
                return True
            except Exception as e:
                logger.error(f"✗ Ошибка сохранения AI кластеров: {e}")
                return False
        
        except Exception as e:
            logger.error(f"✗ Критическая ошибка export_ai_clusters: {e}")
            import traceback
            traceback.print_exc()
            return False